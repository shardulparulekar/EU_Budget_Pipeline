#!/usr/bin/env python3
"""
EU Budget Anomaly Detection Pipeline - Standalone Version
===========================================================

Runs locally with Docker (no Databricks needed)
Tech Stack: PySpark + DuckDB (hybrid architecture maintained)

Data: Place Excel files in data/raw/ folder
Outputs: Results saved to outputs/ folder
"""

import os
import sys
import json
import warnings
from datetime import datetime

# Core imports
import openpyxl
import pandas as pd
import numpy as np
import duckdb

# PySpark - Create local session (instead of using Databricks pre-configured)
from pyspark.sql import SparkSession, functions as F, Row
from pyspark.sql.types import *
from pyspark.sql.window import Window

# ML
from sklearn.ensemble import IsolationForest

# Visualization
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend for Docker
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.ticker as mtick
import seaborn as sns

warnings.filterwarnings('ignore')

# ============================================================================
# SECTION 0: SETUP & CONFIGURATION
# ============================================================================

print("="*70)
print("EU BUDGET ANOMALY DETECTION PIPELINE")
print("Hybrid PySpark + DuckDB Architecture")
print("="*70)

# ============================================================================
# CONFIGURATION - LOCAL FILE SYSTEM (replaces Unity Catalog)
# ============================================================================

# Base directories
if os.path.exists('/app'):  # Running in Docker
    PROJECT_ROOT = '/app'
else:  # Running locally
    PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))

DATA_DIR = os.path.join(PROJECT_ROOT, 'data')
RAW_DATA_PATH = os.path.join(DATA_DIR, 'raw')
OUTPUT_PATH = os.path.join(PROJECT_ROOT, 'outputs')
CHARTS_PATH = os.path.join(OUTPUT_PATH, 'charts')

# Create directories if they don't exist
os.makedirs(RAW_DATA_PATH, exist_ok=True)
os.makedirs(OUTPUT_PATH, exist_ok=True)
os.makedirs(CHARTS_PATH, exist_ok=True)

# Excel files
EXCEL_MAIN = os.path.join(RAW_DATA_PATH, 'eu_budget_spending_and_revenue_2000-2023.xlsx')
EXCEL_SUPP = os.path.join(RAW_DATA_PATH, 'eu_budget_spending_and_revenue_2000-2022.xlsx')

print("\nCONFIGURATION:")
print(f"  Project root: {PROJECT_ROOT}")
print(f"  Raw data: {RAW_DATA_PATH}")
print(f"  Outputs: {OUTPUT_PATH}")

# Verify data file exists
if not os.path.exists(EXCEL_MAIN):
    print(f"\n❌ ERROR: Data file not found!")
    print(f"   Expected: {EXCEL_MAIN}")
    print(f"   Please place Excel file in data/raw/ folder")
    sys.exit(1)
else:
    file_size_mb = os.path.getsize(EXCEL_MAIN) / 1024 / 1024
    print(f"  ✓ Main Excel file: {file_size_mb:.2f} MB")

# ============================================================================
# CREATE LOCAL SPARK SESSION (replaces Databricks pre-configured spark)
# ============================================================================

print("\nInitializing PySpark...")
spark = SparkSession.builder \
    .appName("EU_Budget_Anomaly_Detection") \
    .master("local[*]") \
    .config("spark.driver.memory", "4g") \
    .config("spark.executor.memory", "2g") \
    .config("spark.sql.shuffle.partitions", "8") \
    .config("spark.sql.adaptive.enabled", "true") \
    .config("spark.sql.adaptive.coalescePartitions.enabled", "true") \
    .getOrCreate()

# Set log level to reduce verbosity
spark.sparkContext.setLogLevel("WARN")

print(f"✓ Spark session created")
print(f"  Spark version: {spark.version}")
print(f"  Master: {spark.sparkContext.master}")
print(f"  Cores: {spark.sparkContext.defaultParallelism}")

# ============================================================================
# COUNTRY LISTS & CONSTANTS
# ============================================================================

EU15 = ['BE','DK','DE','EL','ES','FR','IE','IT','LU','NL','AT','PT','FI','SE','UK']
EU25_ADD = ['CY','CZ','EE','HU','LT','LV','MT','PL','SI','SK']
EU27_ADD = ['BG','RO']
EU28_ADD = ['HR']
ALL_CTRY = EU15 + EU25_ADD + EU27_ADD + EU28_ADD

print(f"✓ Countries configured: {len(ALL_CTRY)}")

# ============================================================================
# SECTION 1: DATA INGESTION WITH PYSPARK
# ============================================================================

print("\n" + "="*70)
print("SECTION 1: DATA INGESTION (PySpark)")
print("="*70)

# MFF Heading Crosswalk
MFF_CROSSWALK = {
    '1. agriculture':              'Agriculture & Rural Dev',
    'preservation and management': 'Agriculture & Rural Dev',
    'natural resources':           'Agriculture & Rural Dev',
    '2. structural actions':       'Structural & Cohesion',
    'sustainable growth':          'Structural & Cohesion',
    'smart and inclusive growth':  'Structural & Cohesion',
    'cohesion, resilience':        'Structural & Cohesion',
    '3. internal policies':        'Internal / Single Market',
    'single market, innovation':   'Internal / Single Market',
    '4. external actions':         'External & Global',
    'eu as a global player':       'External & Global',
    'global europe':               'External & Global',
    'neighbourhood and the world': 'External & Global',
    'administration':              'Administration',
    'citizenship, freedom':        'Security & Citizenship',
    'security and citizenship':    'Security & Citizenship',
    'migration and border':        'Security & Citizenship',
    'security and defence':        'Security & Citizenship',
    '7. pre-accession':            'Pre-Accession',
    'nextgenerationeu':            'NGEU Recovery',
}

def map_heading(raw):
    if not raw: return None
    s = str(raw).strip()
    if s.startswith(' '): return None
    low = s.lower()
    for key, val in MFF_CROSSWALK.items():
        if key in low: return val
    return None

print("✓ MFF crosswalk defined")

# Excel Parser Functions
def find_header(rows, after=0):
    """Find row containing country ISO codes."""
    for i, row in enumerate(rows):
        if i < after: continue
        vals = {str(v).strip() for v in row if v is not None}
        if 'BE' in vals and 'DE' in vals and 'Total' in vals:
            return i
    return None

def col_map(header_row, countries):
    """Return {iso: col_index} plus special keys."""
    m = {}
    for j, v in enumerate(header_row):
        s = str(v).strip() if v else ''
        if s in countries:             m[s] = j
        if s == 'Total':               m['__total__'] = j
        if 'earmarked' in s.lower():   m['__earmarked__'] = j
    return m

REV_MAP = {
    'TOTAL national contribution':  'total_contributions',
    'TOTAL National contributions': 'total_contributions',
    'GNP-based own resource':       'gni_contribution',
    'GNI-based own resource':       'gni_contribution',
    'Own resources based on GNI':   'gni_contribution',
    'VAT-based own resource':       'vat_contribution',
    'Own resources based on VAT':   'vat_contribution',
    'Traditional own resources':    'customs',
    'Customs duties':               'customs',
    'Gross National Income':        'gni_size',
    'TOTAL EXPENDITURE':            'total_expenditure',
}

def parse_tab(ws, year, countries):
    """Parse one year-tab → (expenditure_df, revenue_df)."""
    rows = list(ws.iter_rows(values_only=True))
    
    exp_hi = find_header(rows, after=0)
    if exp_hi is None:
        return pd.DataFrame(), pd.DataFrame()
    
    ecm = col_map(list(rows[exp_hi]), countries)
    tc = ecm.get('__total__')
    ec = ecm.get('__earmarked__')
    
    ngeu_start = next(
        (i for i, row in enumerate(rows)
         if any(isinstance(v,str) and 'nextgenerationeu' in v.lower() for v in row)),
        None
    )
    
    exp_rows = []
    for i, row in enumerate(rows[exp_hi+1:], exp_hi+1):
        if any(isinstance(v,str) and v.strip().upper()=='REVENUE' for v in row[:3]):
            break
        
        label = None
        for c in range(4):
            v = row[c] if len(row) > c else None
            if isinstance(v, str) and v.strip():
                candidate = v.strip()
                if map_heading(candidate):
                    label = candidate
                    break
        
        if not label:
            for c in range(3):
                a = row[c] if len(row) > c else None
                b = row[c+1] if len(row) > c+1 else None
                if isinstance(a, str) and isinstance(b, str) and b.strip():
                    if map_heading(b.strip()):
                        label = b.strip()
                        break
        
        if not label: continue
        heading = map_heading(label)
        if not heading: continue
        
        total = row[tc] if tc and tc < len(row) else None
        if not isinstance(total, (int,float)) or total == 0: continue
        
        earmarked = row[ec] if ec and ec < len(row) and isinstance(row[ec],(int,float)) else 0.0
        is_ngeu = 1 if (ngeu_start and i >= ngeu_start) else 0
        
        for ctry in countries:
            col = ecm.get(ctry)
            if col is None: continue
            val = row[col] if col < len(row) else None
            if isinstance(val, (int,float)):
                exp_rows.append({
                    'year':year, 'country':ctry, 'heading_raw':label,
                    'heading':heading, 'actual_meur':round(val,4),
                    'earmarked_meur':round(earmarked,4),
                    'is_ngeu':is_ngeu,
                    'is_preliminary':1 if year==2024 else 0
                })
    
    rev_hi = find_header(rows, after=40) or exp_hi
    rcm = col_map(list(rows[rev_hi]), countries)
    rev_rows = []
    for row in rows[rev_hi+1:]:
        label = None
        for c in range(4):
            v = row[c] if len(row) > c else None
            if isinstance(v, str) and v.strip():
                label = v.strip()
                break
        if not label: continue
        metric = next((v for k,v in REV_MAP.items() if k.lower() in label.lower()), None)
        if not metric: continue
        for ctry in countries:
            col = rcm.get(ctry)
            if col is None: continue
            val = row[col] if col < len(row) else None
            if isinstance(val, (int,float)):
                rev_rows.append({
                    'year':year, 'country':ctry, 
                    'metric':metric, 'value_meur':round(val,4)
                })
    
    return pd.DataFrame(exp_rows), pd.DataFrame(rev_rows)

print("✓ Parser functions defined")

# Parse all Excel sheets
print("\nParsing Excel sheets...")
wb = openpyxl.load_workbook(EXCEL_MAIN)
all_exp, all_rev = [], []

for sheet in wb.sheetnames:
    try: 
        year = int(sheet)
    except: 
        continue
    
    ws = wb[sheet]
    exp_df, rev_df = parse_tab(ws, year, ALL_CTRY)
    
    if exp_df.empty:
        print(f'  ⚠  {year}: no expenditure found')
        continue
    
    all_exp.append(exp_df)
    if not rev_df.empty: 
        all_rev.append(rev_df)
    
    print(f'  ✓  {year}: {len(exp_df):5,} exp  {len(rev_df):5,} rev')

df_exp = pd.concat(all_exp, ignore_index=True)
df_rev = pd.concat(all_rev, ignore_index=True)

print(f'\n✓ Total expenditure rows: {len(df_exp):,}')
print(f'✓ Total revenue rows: {len(df_rev):,}')
print(f'✓ Unified headings: {sorted(df_exp.heading.unique())}')

# Convert to PySpark DataFrames
expenditure_schema = StructType([
    StructField('year', IntegerType(), False),
    StructField('country', StringType(), False),
    StructField('heading_raw', StringType(), True),
    StructField('heading', StringType(), True),
    StructField('actual_meur', DoubleType(), True),
    StructField('earmarked_meur', DoubleType(), True),
    StructField('is_ngeu', IntegerType(), False),
    StructField('is_preliminary', IntegerType(), False)
])

revenue_schema = StructType([
    StructField('year', IntegerType(), False),
    StructField('country', StringType(), False),
    StructField('metric', StringType(), False),
    StructField('value_meur', DoubleType(), True)
])

df_exp_spark = spark.createDataFrame(df_exp, schema=expenditure_schema)
df_rev_spark = spark.createDataFrame(df_rev, schema=revenue_schema)

df_exp_spark.cache()
df_rev_spark.cache()

print(f'\n✓ PySpark DataFrames created')
print(f'  Expenditure: {df_exp_spark.count():,} rows')
print(f'  Revenue: {df_rev_spark.count():,} rows')

# ============================================================================
# DATA QUALITY AUDIT
# ============================================================================

print('\n' + '='*70)
print('COMPREHENSIVE DATA QUALITY AUDIT')
print('='*70)

audit_metrics = {
    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    'pipeline_version': '1.0',
    'data_period': '2000-2024'
}

# 1. Missing Value Analysis
print("\n1. Missing Value Analysis:")
print("-"*40)

missing_counts = df_exp_spark.select([
    F.count(F.when(F.col(c).isNull(), c)).alias(c) 
    for c in df_exp_spark.columns
]).collect()[0]

total_rows = df_exp_spark.count()
exp_missing = {}

for col_name, missing_count in zip(df_exp_spark.columns, missing_counts):
    missing_pct = (missing_count / total_rows) * 100
    status = '✓ OK' if missing_pct < 20 else '⚠ WARNING'
    
    exp_missing[col_name] = {
        'count': missing_count,
        'pct': round(missing_pct, 2),
        'status': status
    }
    
    if missing_count > 0:
        print(f'   {col_name}: {missing_count:,} ({missing_pct:.2f}%) - {status}')

if sum(m['count'] for m in exp_missing.values()) == 0:
    print('   ✓ No missing values detected')

audit_metrics['missing_values'] = exp_missing

# 2. Duplicate Detection
print('\n2. Duplicate Detection:')
print("-"*40)

distinct_records = df_exp_spark.distinct().count()
duplicates = total_rows - distinct_records
duplicate_pct = (duplicates / total_rows) * 100 if total_rows > 0 else 0

print(f'   Total records: {total_rows:,}')
print(f'   Distinct records: {distinct_records:,}')
print(f'   Duplicates: {duplicates:,} ({duplicate_pct:.2f}%)')
print(f'   Status: {"✓ OK - No duplicates" if duplicates == 0 else "⚠ Review duplicates"}')

audit_metrics['duplicates'] = {
    'total': total_rows,
    'distinct': distinct_records,
    'duplicate_count': duplicates,
    'duplicate_pct': round(duplicate_pct, 2),
    'status': '✓ OK' if duplicates == 0 else '⚠ Review'
}

# 3. Outlier Analysis
print('\n3. Outlier Detection (actual_meur):')
print("-"*40)

amount_stats = df_exp_spark.select(
    F.mean('actual_meur').alias('mean'),
    F.stddev('actual_meur').alias('stddev'),
    F.min('actual_meur').alias('min'),
    F.max('actual_meur').alias('max'),
    F.expr('percentile_approx(actual_meur, 0.25)').alias('q25'),
    F.expr('percentile_approx(actual_meur, 0.5)').alias('median'),
    F.expr('percentile_approx(actual_meur, 0.75)').alias('q75')
).collect()[0]

iqr = amount_stats.q75 - amount_stats.q25
lower_bound = amount_stats.q25 - 3 * iqr
upper_bound = amount_stats.q75 + 3 * iqr

outliers = df_exp_spark.filter(
    (F.col('actual_meur') < lower_bound) | 
    (F.col('actual_meur') > upper_bound)
).count()
outlier_pct = (outliers / total_rows) * 100

print(f'   Mean: €{amount_stats.mean:,.2f}M')
print(f'   Median: €{amount_stats.median:,.2f}M')
print(f'   Std Dev: €{amount_stats.stddev:,.2f}M')
print(f'   Range: €{amount_stats["min"]:,.2f}M - €{amount_stats["max"]:,.2f}M')
print(f'   IQR: €{iqr:,.2f}M')
print(f'   Outliers (>3 IQR): {outliers:,} ({outlier_pct:.2f}%)')

audit_metrics['outliers'] = {
    'mean': round(amount_stats.mean, 2),
    'median': round(amount_stats.median, 2),
    'std': round(amount_stats.stddev, 2),
    'min': round(amount_stats['min'], 2),
    'max': round(amount_stats['max'], 2),
    'iqr': round(iqr, 2),
    'outlier_count': outliers,
    'outlier_pct': round(outlier_pct, 2)
}

# 4-8: Additional quality checks...
# (Country coverage, heading dist, temporal dist, business rules, overall score)
# [Including complete code from merged audit cell - abbreviated here for space]

country_coverage = df_exp_spark.groupBy('country').agg(
    F.min('year').alias('first_year'),
    F.max('year').alias('last_year'),
    F.countDistinct('year').alias('years_present')
).toPandas()

country_coverage['coverage_pct'] = (country_coverage['years_present'] / 25) * 100
unique_countries = len(country_coverage)
full_coverage = (country_coverage['coverage_pct'] == 100).sum()

print(f'\n4. Country Coverage: {unique_countries} countries, {full_coverage} with full coverage')

heading_dist = df_exp_spark.groupBy('heading').agg(
    F.count('*').alias('record_count'),
    F.avg('actual_meur').alias('avg_amount')
).toPandas().sort_values('record_count', ascending=False)

print(f'5. Heading Distribution: {len(heading_dist)} headings')

temporal_dist = df_exp_spark.groupBy('year').agg(
    F.count('*').alias('record_count')
).toPandas().sort_values('year')

print(f'6. Temporal Distribution: {len(temporal_dist)} years')

negative_amounts = df_exp_spark.filter(F.col('actual_meur') < 0).count()
print(f'7. Business Rules: {negative_amounts} negative amounts')

# Overall quality score
completeness_score = 100 - (sum(m['pct'] for m in exp_missing.values() if m['pct'] > 0) / len(exp_missing))
accuracy_score = 100 - (outlier_pct / 2)
consistency_score = 100 - duplicate_pct
timeliness_score = 100 if temporal_dist['year'].max() >= 2023 else 90

overall_score = (completeness_score + accuracy_score + consistency_score + timeliness_score) / 4
status = '✓ EXCELLENT' if overall_score >= 95 else '✓ GOOD' if overall_score >= 85 else '⚠ NEEDS IMPROVEMENT'

print(f'\n8. Overall Quality Score: {overall_score:.1f}% ({status})')

audit_metrics['quality_scores'] = {
    'completeness': round(completeness_score, 1),
    'accuracy': round(accuracy_score, 1),
    'consistency': round(consistency_score, 1),
    'timeliness': round(timeliness_score, 1),
    'overall': round(overall_score, 1),
    'status': status
}

# Save audit report (local file system, not Unity Catalog)
audit_report_path = os.path.join(OUTPUT_PATH, 'data_quality_audit.json')
with open(audit_report_path, 'w') as f:
    json.dump(audit_metrics, f, indent=2, default=str)

print(f'\n✓ Audit report saved to: {audit_report_path}')

# ============================================================================
# SECTION 2: FEATURE ENGINEERING WITH DUCKDB
# ============================================================================

print('\n' + '='*70)
print('SECTION 2: FEATURE ENGINEERING (DuckDB)')
print('='*70)

# Export PySpark to DuckDB
df_exp = df_exp_spark.toPandas()
df_rev = df_rev_spark.toPandas()

# Pivot revenue
df_rev_wide = df_rev.pivot_table(
    index=['year','country'], 
    columns='metric',
    values='value_meur', 
    aggfunc='first'
).reset_index()

df_rev_wide.columns.name = None
for col in ['total_contributions','gni_size','gni_contribution',
            'vat_contribution','customs','total_expenditure']:
    if col not in df_rev_wide.columns:
        df_rev_wide[col] = np.nan

# Initialize DuckDB
conn = duckdb.connect(':memory:')
conn.register('exp_raw', df_exp)
conn.register('rev_wide', df_rev_wide)

print("✓ Data registered with DuckDB")
print(f"  Expenditure: {len(df_exp):,} rows")
print(f"  Revenue: {len(df_rev_wide):,} rows")

# Create intermediate tables
conn.execute("""
    CREATE OR REPLACE TABLE heading_level AS
    SELECT year, country, heading,
           SUM(actual_meur) AS actual_meur,
           SUM(earmarked_meur) AS earmarked_meur,
           MAX(is_ngeu) AS is_ngeu,
           MAX(is_preliminary) AS is_preliminary
    FROM exp_raw
    GROUP BY year, country, heading
""")

conn.execute("""
    CREATE OR REPLACE TABLE total_receipts AS
    SELECT year, country, SUM(actual_meur) AS receipts_meur
    FROM exp_raw
    WHERE is_ngeu = 0
    GROUP BY year, country
""")

print("✓ Intermediate tables created")

# Master feature table with 22 features
conn.execute("""
    CREATE OR REPLACE TABLE features AS
    
    WITH base AS (
        SELECT
            h.year, h.country, h.heading,
            h.actual_meur, h.earmarked_meur,
            h.is_ngeu, h.is_preliminary,
            
            t.receipts_meur,
            r.total_contributions, r.gni_size,
            ROUND((t.receipts_meur - r.total_contributions)
                  / NULLIF(r.gni_size, 0) * 100, 4) AS net_pct_gni,
            ROUND(t.receipts_meur - r.total_contributions, 2) AS net_position_meur,
            
            ROUND(h.earmarked_meur - h.actual_meur, 2) AS absorption_gap_meur,
            ROUND(100.0 * h.actual_meur / NULLIF(h.earmarked_meur, 0), 2) AS absorption_rate_pct,
            
            ROUND(h.actual_meur - LAG(h.actual_meur) OVER
                (PARTITION BY h.country, h.heading ORDER BY h.year), 2) AS yoy_change,
            
            LAG(h.actual_meur,1) OVER (PARTITION BY h.country, h.heading ORDER BY h.year) AS lag1,
            LAG(h.actual_meur,2) OVER (PARTITION BY h.country, h.heading ORDER BY h.year) AS lag2,
            LAG(h.actual_meur,3) OVER (PARTITION BY h.country, h.heading ORDER BY h.year) AS lag3,
            
            ROUND(AVG(h.actual_meur) OVER (
                PARTITION BY h.country, h.heading ORDER BY h.year
                ROWS BETWEEN 3 PRECEDING AND 1 PRECEDING
            ), 2) AS roll_mean_3,
            
            h.year - 2000 AS year_index
            
        FROM heading_level h
        LEFT JOIN rev_wide r ON h.year = r.year AND h.country = r.country
        LEFT JOIN total_receipts t ON h.year = t.year AND h.country = t.country
    )
    
    SELECT *,
        ROUND(
            (absorption_gap_meur - AVG(absorption_gap_meur) OVER (PARTITION BY country, heading))
            / NULLIF(STDDEV(absorption_gap_meur) OVER (PARTITION BY country, heading), 0),
        3) AS zscore_absorption,
        
        ROUND(
            (net_pct_gni - AVG(net_pct_gni) OVER (PARTITION BY country))
            / NULLIF(STDDEV(net_pct_gni) OVER (PARTITION BY country), 0),
        3) AS zscore_net
    
    FROM base
    ORDER BY country, heading, year
""")

df_feat = conn.execute('SELECT * FROM features').df()
conn.close()

print(f'✓ Feature table created: {len(df_feat):,} rows × {len(df_feat.columns)} columns')

# ============================================================================
# SECTION 3: H1 NET POSITION ANALYSIS (PySpark)
# ============================================================================

print('\n' + '='*70)
print('SECTION 3: H1 NET POSITION ANALYSIS (PySpark)')
print('='*70)

df_feat_spark = spark.createDataFrame(df_feat)

df_h1_spark = df_feat_spark.groupBy('country', 'year') \
    .agg(
        F.first('net_position_meur').alias('net_position_meur'),
        F.first('net_pct_gni').alias('net_pct_gni'),
        F.first('zscore_net').alias('zscore_net'),
        F.first('receipts_meur').alias('receipts_meur'),
        F.first('total_contributions').alias('total_contributions')
    ) \
    .withColumn('h1_anomaly', F.when(F.abs(F.col('zscore_net')) > 2.0, 1).otherwise(0))

country_window = Window.partitionBy('country').orderBy('year')

df_h1_spark = df_h1_spark \
    .withColumn('net_position_lag1', F.lag('net_position_meur', 1).over(country_window)) \
    .withColumn('net_position_change', F.col('net_position_meur') - F.col('net_position_lag1'))

h1_summary = df_h1_spark.agg(
    F.count('*').alias('country_years'),
    F.sum('h1_anomaly').alias('total_anomalies'),
    F.countDistinct('country').alias('unique_countries')
).collect()[0]

print(f'  Country-years analyzed: {h1_summary.country_years:,}')
print(f'  Anomalies detected: {h1_summary.total_anomalies}')
print(f'  Countries covered: {h1_summary.unique_countries}')

# ============================================================================
# SECTION 4: H3 ABSORPTION ANOMALY DETECTION (PySpark)
# ============================================================================

print('\n' + '='*70)
print('SECTION 4: H3 ABSORPTION ANOMALY DETECTION (PySpark)')
print('='*70)

df_h3 = df_feat_spark.filter(
    (F.col('is_ngeu') == 0) & 
    (F.col('is_preliminary') == 0) &
    F.col('lag1').isNotNull()
)

df_h3_pd = df_h3.select(
    'year', 'country', 'heading',
    'absorption_gap_meur', 'absorption_rate_pct',
    'lag1', 'lag2', 'lag3', 'roll_mean_3',
    'zscore_absorption'
).toPandas().fillna({'absorption_rate_pct': 0.0})

print(f'✓ H3 working set: {len(df_h3_pd):,} rows')

# Isolation Forest (Primary Method)
feature_cols = ['absorption_gap_meur', 'absorption_rate_pct', 'lag1', 'roll_mean_3']
X = df_h3_pd[feature_cols].fillna(0)

print("Running Isolation Forest...")
iso_forest = IsolationForest(
    contamination=0.05,
    random_state=42,
    n_estimators=100
)
iso_forest.fit(X)

df_h3_pd['if_anomaly'] = (iso_forest.predict(X) == -1).astype(int)
df_h3_pd['if_score'] = iso_forest.score_samples(X)

# Local Outlier Factor (Baseline - Better than Z-score)
print("Running Local Outlier Factor (baseline)...")
from sklearn.neighbors import LocalOutlierFactor

lof = LocalOutlierFactor(
    n_neighbors=20,  # Consider 20 nearest neighbors
    contamination=0.05,  # Same as Isolation Forest for comparison
    novelty=False  # Fit-predict mode
)

df_h3_pd['lof_anomaly'] = (lof.fit_predict(X) == -1).astype(int)
df_h3_pd['lof_score'] = lof.negative_outlier_factor_  # Higher = more normal

# Compare methods
n_total = len(df_h3_pd)
n_if = df_h3_pd['if_anomaly'].sum()
n_lof = df_h3_pd['lof_anomaly'].sum()
n_both = ((df_h3_pd['if_anomaly'] == 1) & (df_h3_pd['lof_anomaly'] == 1)).sum()

print(f'  Total observations: {n_total:,}')
print(f'  Isolation Forest anomalies: {n_if} ({n_if/n_total*100:.1f}%)')
print(f'  LOF anomalies (baseline): {n_lof} ({n_lof/n_total*100:.1f}%)')
print(f'  Both methods agree: {n_both} ({n_both/max(n_if,n_lof)*100:.1f}% agreement)')
print(f'  Precision (IF vs LOF): {n_both/n_if*100:.1f}%' if n_if > 0 else '  Precision: N/A')

# ============================================================================
# SECTION 5: EXPORT RESULTS (Local File System)
# ============================================================================

print('\n' + '='*70)
print('SECTION 5: EXPORT RESULTS')
print('='*70)

# Save outputs (using local file system instead of Unity Catalog)
print(f"Saving results to: {OUTPUT_PATH}")

# Feature table
df_feat.to_parquet(os.path.join(OUTPUT_PATH, 'eu_features.parquet'), index=False)
print(f'✓ eu_features.parquet')

# H1 results
df_h1_spark.toPandas().to_csv(os.path.join(OUTPUT_PATH, 'h1_net_positions.csv'), index=False)
print(f'✓ h1_net_positions.csv')

# H3 results
df_h3_pd.to_parquet(os.path.join(OUTPUT_PATH, 'h3_anomaly_scores.parquet'), index=False)
print(f'✓ h3_anomaly_scores.parquet')

# Top anomalies (from Isolation Forest)
top = (
    df_h3_pd[df_h3_pd['if_anomaly'] == 1]
    .sort_values('if_score', ascending=True)  # Most anomalous first (lower score = more anomalous)
    [['year','country','heading','absorption_gap_meur','absorption_rate_pct',
      'if_score','lof_anomaly','lof_score']]
    .head(30)
    .reset_index(drop=True)
)
top.to_csv(os.path.join(OUTPUT_PATH, 'anomaly_ranked.csv'), index=False)
print(f'✓ anomaly_ranked.csv')

# Summary JSON
summary = {
    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    'feature_rows': len(df_feat),
    'h1_country_years': int(h1_summary.country_years),
    'h1_anomalies': int(h1_summary.total_anomalies),
    'h3_rows_scored': n_total,
    'h3_if_anomalies': int(n_if),
    'h3_lof_anomalies': int(n_lof),
    'h3_both_agree': int(n_both),
    'agreement_rate_pct': round(n_both / max(n_if, 1) * 100, 1),
    'precision_if_vs_lof': round(n_both / n_if * 100, 1) if n_if > 0 else 0,
    'baseline_method': 'Local Outlier Factor (LOF)',
    'overall_quality_score': audit_metrics['quality_scores']['overall']
}

with open(os.path.join(OUTPUT_PATH, 'pipeline_summary.json'), 'w') as f:
    json.dump(summary, f, indent=2)
print(f'✓ pipeline_summary.json')

print('\n' + '='*70)
print('PIPELINE COMPLETE')
print('='*70)
print(json.dumps(summary, indent=2))
print(f'\nAll outputs saved to: {OUTPUT_PATH}')

# ============================================================================
# SECTION 6: EDA VISUALIZATIONS
# ============================================================================

print('\n' + '='*70)
print('SECTION 6: EDA VISUALIZATIONS')
print('='*70)

# Convert to pandas for visualization (lighter weight)
df_feat_viz = df_feat_spark.toPandas()

# Chart A: Row counts per year
plt.figure(figsize=(14, 6))
yearly_counts = df_feat_viz.groupby('year').size().reset_index(name='count')
plt.bar(yearly_counts['year'], yearly_counts['count'], color='steelblue', alpha=0.7)
plt.axvline(2007, color='red', linestyle='--', alpha=0.5, label='MFF 2007-2013')
plt.axvline(2014, color='orange', linestyle='--', alpha=0.5, label='MFF 2014-2020')
plt.axvline(2021, color='green', linestyle='--', alpha=0.5, label='MFF 2021-2027')
plt.xlabel('Year', fontsize=12)
plt.ylabel('Number of Records', fontsize=12)
plt.title('EU Budget Records by Year (MFF Transitions)', fontsize=14, fontweight='bold')
plt.legend()
plt.grid(axis='y', alpha=0.3)
plt.tight_layout()
plt.savefig(os.path.join(CHARTS_PATH, 'eda_a_row_counts.png'), dpi=150, bbox_inches='tight')
plt.close()
print('✓ Chart A: Row counts by year')

# Chart B: Country coverage heatmap
plt.figure(figsize=(16, 10))
coverage_pivot = df_feat_viz.groupby(['country', 'year']).size().reset_index(name='records')
coverage_pivot = coverage_pivot.pivot(index='country', columns='year', values='records').fillna(0)
sns.heatmap(coverage_pivot, cmap='YlGnBu', cbar_kws={'label': 'Records'}, linewidths=0.5)
plt.title('Country Coverage Heatmap (2000-2024)', fontsize=14, fontweight='bold')
plt.xlabel('Year', fontsize=12)
plt.ylabel('Country', fontsize=12)
plt.tight_layout()
plt.savefig(os.path.join(CHARTS_PATH, 'eda_b_country_coverage.png'), dpi=150, bbox_inches='tight')
plt.close()
print('✓ Chart B: Country coverage heatmap')

# Chart C: Heading distribution
plt.figure(figsize=(12, 7))
heading_counts = df_feat_viz['heading'].value_counts()
plt.barh(range(len(heading_counts)), heading_counts.values, color='teal', alpha=0.7)
plt.yticks(range(len(heading_counts)), heading_counts.index, fontsize=10)
plt.xlabel('Number of Records', fontsize=12)
plt.title('Distribution of Records by Budget Heading', fontsize=14, fontweight='bold')
plt.grid(axis='x', alpha=0.3)
plt.tight_layout()
plt.savefig(os.path.join(CHARTS_PATH, 'eda_c_heading_distribution.png'), dpi=150, bbox_inches='tight')
plt.close()
print('✓ Chart C: Heading distribution')

# Chart D: Total budget over time
plt.figure(figsize=(14, 6))
total_by_year = df_feat_viz.groupby('year')['actual_meur'].sum() / 1000  # Convert to billions
plt.plot(total_by_year.index, total_by_year.values, marker='o', linewidth=2, markersize=6, color='darkblue')
plt.fill_between(total_by_year.index, total_by_year.values, alpha=0.3, color='skyblue')
plt.xlabel('Year', fontsize=12)
plt.ylabel('Total Budget (€ Billions)', fontsize=12)
plt.title('Total EU Budget Expenditure Over Time', fontsize=14, fontweight='bold')
plt.grid(alpha=0.3)
plt.tight_layout()
plt.savefig(os.path.join(CHARTS_PATH, 'eda_d_total_budget.png'), dpi=150, bbox_inches='tight')
plt.close()
print('✓ Chart D: Total budget over time')

# ============================================================================
# SECTION 7: ANOMALY VISUALIZATIONS
# ============================================================================

print('\n' + '='*70)
print('SECTION 7: ANOMALY VISUALIZATIONS')
print('='*70)

# Prepare H1 and H3 data
df_h1_viz = df_h1_spark.toPandas()

# H1-1: Net position drift (top countries)
plt.figure(figsize=(14, 8))
top_countries = ['DE', 'FR', 'IT', 'PL', 'ES', 'NL']
for country in top_countries:
    country_data = df_h1_viz[df_h1_viz['country'] == country].sort_values('year')
    plt.plot(country_data['year'], country_data['net_position_meur'] / 1000, 
             marker='o', label=country, linewidth=2)

plt.axhline(0, color='black', linestyle='--', alpha=0.5)
plt.xlabel('Year', fontsize=12)
plt.ylabel('Net Position (€ Billions)', fontsize=12)
plt.title('H1: Net Position Drift - Top Contributors & Receivers', fontsize=14, fontweight='bold')
plt.legend(loc='best', fontsize=10)
plt.grid(alpha=0.3)
plt.tight_layout()
plt.savefig(os.path.join(CHARTS_PATH, 'h1_net_position_drift.png'), dpi=150, bbox_inches='tight')
plt.close()
print('✓ H1-1: Net position drift')

# H1-2: Contributors vs Receivers (latest year)
plt.figure(figsize=(12, 10))
latest_year = df_h1_viz['year'].max()
latest_data = df_h1_viz[df_h1_viz['year'] == latest_year].sort_values('net_position_meur')
colors = ['red' if x < 0 else 'green' for x in latest_data['net_position_meur']]
plt.barh(range(len(latest_data)), latest_data['net_position_meur'] / 1000, color=colors, alpha=0.7)
plt.yticks(range(len(latest_data)), latest_data['country'], fontsize=9)
plt.axvline(0, color='black', linestyle='-', linewidth=1)
plt.xlabel('Net Position (€ Billions)', fontsize=12)
plt.title(f'H1: Net Contributors vs Receivers ({latest_year})', fontsize=14, fontweight='bold')
plt.grid(axis='x', alpha=0.3)
plt.tight_layout()
plt.savefig(os.path.join(CHARTS_PATH, 'h1_contributors_receivers.png'), dpi=150, bbox_inches='tight')
plt.close()
print('✓ H1-2: Contributors vs receivers')

# H3-1: Anomaly heatmap
plt.figure(figsize=(16, 10))
anomaly_pivot = df_h3_pd.groupby(['country', 'year'])['if_anomaly'].sum().reset_index()
anomaly_pivot = anomaly_pivot.pivot(index='country', columns='year', values='if_anomaly').fillna(0)
sns.heatmap(anomaly_pivot, cmap='YlOrRd', cbar_kws={'label': 'Anomaly Count'}, linewidths=0.5)
plt.title('H3: Absorption Anomalies by Country and Year', fontsize=14, fontweight='bold')
plt.xlabel('Year', fontsize=12)
plt.ylabel('Country', fontsize=12)
plt.tight_layout()
plt.savefig(os.path.join(CHARTS_PATH, 'h3_anomaly_heatmap.png'), dpi=150, bbox_inches='tight')
plt.close()
print('✓ H3-1: Anomaly heatmap')

# H3-2: Absorption rate by heading
plt.figure(figsize=(12, 7))
heading_absorption = df_h3_pd.groupby('heading')['absorption_rate_pct'].mean().sort_values()
plt.barh(range(len(heading_absorption)), heading_absorption.values, color='purple', alpha=0.7)
plt.yticks(range(len(heading_absorption)), heading_absorption.index, fontsize=10)
plt.axvline(100, color='red', linestyle='--', alpha=0.5, label='100% (Full Absorption)')
plt.xlabel('Average Absorption Rate (%)', fontsize=12)
plt.title('H3: Average Absorption Rate by Budget Heading', fontsize=14, fontweight='bold')
plt.legend()
plt.grid(axis='x', alpha=0.3)
plt.tight_layout()
plt.savefig(os.path.join(CHARTS_PATH, 'h3_absorption_by_heading.png'), dpi=150, bbox_inches='tight')
plt.close()
print('✓ H3-2: Absorption by heading')

# H3-3: Top 10 anomalies
plt.figure(figsize=(12, 8))
top_10 = df_h3_pd.nlargest(10, 'if_score')[['year', 'country', 'heading', 'if_score']].reset_index(drop=True)
top_10['label'] = top_10.apply(lambda x: f"{x['country']} {x['year']}\n{x['heading'][:20]}", axis=1)
plt.barh(range(len(top_10)), top_10['if_score'], color='crimson', alpha=0.7)
plt.yticks(range(len(top_10)), top_10['label'], fontsize=9)
plt.xlabel('Isolation Forest Anomaly Score', fontsize=12)
plt.title('H3: Top 10 Absorption Anomalies', fontsize=14, fontweight='bold')
plt.grid(axis='x', alpha=0.3)
plt.tight_layout()
plt.savefig(os.path.join(CHARTS_PATH, 'h3_top_10_anomalies.png'), dpi=150, bbox_inches='tight')
plt.close()
print('✓ H3-3: Top 10 anomalies')

# H3-4: Policy events overlay
plt.figure(figsize=(14, 6))
anomalies_by_year = df_h3_pd.groupby('year')['if_anomaly'].sum()
plt.plot(anomalies_by_year.index, anomalies_by_year.values, marker='o', linewidth=2, 
         markersize=8, color='darkred', label='Anomaly Count')
plt.axvline(2004, color='blue', linestyle='--', alpha=0.5, label='EU Enlargement (2004)')
plt.axvline(2008, color='orange', linestyle='--', alpha=0.5, label='Financial Crisis (2008)')
plt.axvline(2020, color='green', linestyle='--', alpha=0.5, label='COVID-19 (2020)')
plt.xlabel('Year', fontsize=12)
plt.ylabel('Number of Anomalies', fontsize=12)
plt.title('H3: Anomalies Over Time (Policy Events Overlay)', fontsize=14, fontweight='bold')
plt.legend(loc='upper left')
plt.grid(alpha=0.3)
plt.tight_layout()
plt.savefig(os.path.join(CHARTS_PATH, 'h3_policy_events.png'), dpi=150, bbox_inches='tight')
plt.close()
print('✓ H3-4: Policy events')

# H3-5: Chronic underabsorbers
plt.figure(figsize=(16, 10))
chronic_countries = ['ES', 'IT', 'PL', 'RO']
fig, axes = plt.subplots(2, 2, figsize=(16, 10))
for idx, country in enumerate(chronic_countries):
    ax = axes[idx // 2, idx % 2]
    country_data = df_h3_pd[df_h3_pd['country'] == country]
    
    ax.scatter(country_data[country_data['if_anomaly'] == 0]['year'],
               country_data[country_data['if_anomaly'] == 0]['absorption_rate_pct'],
               alpha=0.5, s=30, label='Normal', color='green')
    ax.scatter(country_data[country_data['if_anomaly'] == 1]['year'],
               country_data[country_data['if_anomaly'] == 1]['absorption_rate_pct'],
               alpha=0.8, s=80, label='Anomaly', color='red', marker='X')
    
    ax.axhline(100, color='black', linestyle='--', alpha=0.3)
    ax.set_xlabel('Year', fontsize=10)
    ax.set_ylabel('Absorption Rate (%)', fontsize=10)
    ax.set_title(f'{country} - Absorption Anomalies', fontsize=12, fontweight='bold')
    ax.legend()
    ax.grid(alpha=0.3)

plt.tight_layout()
plt.savefig(os.path.join(CHARTS_PATH, 'h3_chronic_underabsorbers.png'), dpi=150, bbox_inches='tight')
plt.close()
print('✓ H3-5: Chronic underabsorbers')

print(f'\n✓ All 11 charts saved to: {CHARTS_PATH}')

# ============================================================================
# FINAL SUMMARY
# ============================================================================

print('\n' + '='*70)
print('PIPELINE EXECUTION COMPLETE')
print('='*70)
print(json.dumps(summary, indent=2))
print(f'\nOutputs saved to: {OUTPUT_PATH}')
print(f'Charts saved to: {CHARTS_PATH}')
print(f'\nQuality Score: {overall_score:.1f}% ({status})')
print('='*70)

# Stop Spark session
spark.stop()
print('\n✓ Spark session stopped')
